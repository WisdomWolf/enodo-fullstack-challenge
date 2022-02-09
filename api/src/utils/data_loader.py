import logging

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from models.real_estate import RealEstateProperty
from utils.directory_resolvers import get_db_uri, get_file_path

logger = logging.getLogger(__name__)


def resolve_headers(worksheet):
    headers = [
        cell.value.lower().replace(' ', '_')
        for cell in worksheet['1']
    ]
    return headers


def convert_excel_to_db_objects(filename, entity_class):
    wb = load_workbook(filename)
    ws = wb.worksheets[0]
    headers = resolve_headers(ws)
    database_objects = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        kwargs = {
            key: cell
            for key, cell in zip(headers, row)
        }
        database_objects.append(entity_class(**kwargs))
    return database_objects


def insert_data(engine, database_objects):
    # TODO: Make idempotent to prevent data duplication
    Session = sessionmaker(engine)
    with Session() as session:
        try:
            session.add_all(database_objects)
            session.commit()
        except Exception:
            logger.exception('Something went wrong when loading data to database')


if __name__ == '__main__':
    excel_file_path = get_file_path('Enodo_Skills_Assessment_Data_File.xlsx')
    db_objects = convert_excel_to_db_objects(excel_file_path, RealEstateProperty)

    db_uri = get_db_uri('real_estate.db')
    engine = create_engine(db_uri)
    insert_data(engine, db_objects)
    print('Data insertion completed successfully')
