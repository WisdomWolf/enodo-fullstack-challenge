from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from data.models import RealEstateProperty


def resolve_headers(worksheet):
    headers = [
        cell.value.lower().replace(' ', '_')
        for cell in worksheet['1']
    ]
    return headers


def convert_excel_to_db_objects(filename, entity_class):
    wb = load_workbook(filename)
    ws = wb.worksheets[0]
    headers = resolve_headers(ws)
    database_objects = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        kwargs = {
            key: cell
            for key, cell in zip(headers, row)
        }
        database_objects.append(entity_class(**kwargs))
    return database_objects


def insert_data(engine, database_objects):
    # TODO: Make idempotent to prevent data duplication
    Session = sessionmaker(engine)
    with Session() as session:
        session.add_all(database_objects)
        session.commit()


if __name__ == '__main__':
    project_path = Path(__file__).parents[3]
    excel_file_path = project_path.joinpath(
        'source_data/Enodo_Skills_Assessment_Data_File.xlsx')
    db_objects = convert_excel_to_db_objects(excel_file_path, RealEstateProperty)

    db_path = project_path.joinpath('api/data/real_estate.db')
    engine = create_engine(f'sqlite:///{db_path}')
    insert_data(engine, db_objects)
